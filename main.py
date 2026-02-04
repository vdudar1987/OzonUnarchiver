import requests
import json
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from threading import Thread
from openpyxl import load_workbook, Workbook
import traceback
from queue import Queue
import time

CONFIG_FILE = "config.json"
API_BASE = "https://api-seller.ozon.ru"


def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def get_headers(account):
    return {
        "Client-Id": account["client_id"],
        "Api-Key": account["api_key"],
        "Content-Type": "application/json"
    }


def read_offer_ids_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # ищем колонку offer_id
    first_row = [cell.value for cell in ws[1]]
    headers = [str(h).strip().lower() if h is not None else "" for h in first_row]
    
    if "offer_id" in headers:
        col_idx = headers.index("offer_id") + 1
        start_row = 2
    else:
        col_idx = 1
        start_row = 1

    offer_ids = []
    for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx, values_only=True):
        v = row[0]
        if v is None:
            continue
        offer_ids.append(str(v).strip())
    
    # Удаляем пустые значения
    offer_ids = [oid for oid in offer_ids if oid and oid != ""]
    return offer_ids


def write_report_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["offer_id", "product_id", "status_before", "archived", "is_autoarchived", "action", "error"])
    for r in rows:
        ws.append([
            r.get("offer_id"), 
            r.get("product_id"), 
            r.get("status_before"),
            r.get("archived"),
            r.get("is_autoarchived"),
            r.get("action"),
            r.get("error", "")
        ])
    wb.save(path)


def get_products_info(account, offer_ids, log_callback):
    url = f"{API_BASE}/v3/product/info/list"
    headers = get_headers(account)
    results = []
    errors = []

    log_callback(f"Начинаем проверку {len(offer_ids)} товаров...")

    for i in range(0, len(offer_ids), 100):
        chunk = offer_ids[i:i + 100]
        payload = {
            "offer_id": chunk,
            "product_id": []
        }
        
        try:
            log_callback(f"Запрос товаров {i+1}-{min(i+100, len(offer_ids))}...")
            
            resp = requests.post(url, headers=headers, json=payload, timeout=30)
            
            if resp.status_code != 200:
                error_msg = f"Ошибка API: {resp.status_code}"
                log_callback(error_msg)
                errors.append(error_msg)
                continue
            
            response_data = resp.json()

            items = response_data.get("items")
            if items is None:
                items = response_data.get("result", {}).get("items")

            if items is not None:
                batch = items
                results.extend(batch)
                log_callback(f"Получено {len(batch)} товаров")
                # Для отладки - посмотрим структуру первого товара
                if batch and i == 0:
                    first_item_keys = list(batch[0].keys())
                    log_callback(f"Ключи первого товара: {first_item_keys}")
            else:
                response_keys = list(response_data.keys())
                log_callback(f"Нет поля 'items' в ответе. Ключи ответа: {response_keys}")
                errors.append(f"Нет поля 'items' в ответе. Ключи ответа: {response_keys}")
            
        except requests.exceptions.RequestException as e:
            error_msg = f"Ошибка запроса: {str(e)}"
            log_callback(error_msg)
            errors.append(error_msg)
        except json.JSONDecodeError as e:
            error_msg = f"Ошибка парсинга JSON: {str(e)}"
            log_callback(error_msg)
            errors.append(error_msg)
        except Exception as e:
            error_msg = f"Неожиданная ошибка: {str(e)}"
            log_callback(error_msg)
            errors.append(error_msg)
        
        log_callback(f"Проверено {len(results)} из {len(offer_ids)} товаров")
    
    if errors:
        log_callback(f"Обнаружены ошибки: {len(errors)}")
    
    return results, errors


def unarchive_products(account, product_ids, log_callback):
    url = f"{API_BASE}/v1/product/unarchive"
    headers = get_headers(account)
    
    # Разбиваем на батches по 100 товаров
    results = []
    errors = []
    
    for i in range(0, len(product_ids), 100):
        chunk = product_ids[i:i + 100]
        payload = {"product_id": chunk}
        
        try:
            log_callback(f"Восстанавливаем товары {i+1}-{min(i+100, len(product_ids))}...")
            resp = requests.post(url, headers=headers, json=payload, timeout=30)
            
            if resp.status_code != 200:
                error_msg = f"Ошибка восстановления: {resp.status_code}"
                log_callback(error_msg)
                errors.append(error_msg)
                continue
            
            result = resp.json()
            results.append(result)
            log_callback(f"Батч восстановлен")
            
        except Exception as e:
            error_msg = f"Ошибка при восстановлении батча: {str(e)}"
            log_callback(error_msg)
            errors.append(error_msg)
    
    return results, errors


class OzonUnarchiverApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Ozon Unarchiver v2.0")
        self.root.geometry("700x500")
        self.root.resizable(True, True)

        if not os.path.exists(CONFIG_FILE):
            messagebox.showerror("Ошибка", f"Файл {CONFIG_FILE} не найден.")
            root.destroy()
            return

        try:
            self.config = load_config()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить конфигурацию: {e}")
            root.destroy()
            return

        self.file_path = None
        self.log_queue = Queue()
        
        # UI
        tk.Label(root, text="Выберите магазин:", font=("Arial", 12)).pack(pady=6)
        self.account_var = tk.StringVar()
        self.account_menu = ttk.Combobox(
            root, textvariable=self.account_var, state="readonly",
            values=[acc["name"] for acc in self.config["accounts"]]
        )
        self.account_menu.pack(pady=5, ipadx=5)
        if self.config["accounts"]:
            self.account_menu.current(0)

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=8)
        tk.Button(btn_frame, text="Выбрать файл (.xlsx)", command=self.select_file).grid(row=0, column=0, padx=6)
        tk.Button(btn_frame, text="Запустить", command=self.run_process).grid(row=0, column=1, padx=6)
        tk.Button(btn_frame, text="Тест API", command=self.test_api).grid(row=0, column=2, padx=6)

        self.file_label = tk.Label(root, text="Файл не выбран", fg="gray")
        self.file_label.pack()

        # Прогресс бар
        self.progress = ttk.Progressbar(root, mode='determinate')
        self.progress.pack(pady=5, fill='x', padx=20)

        # Скроллбар для лог бокса
        log_frame = tk.Frame(root)
        log_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        scrollbar = tk.Scrollbar(log_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.log_box = tk.Text(log_frame, height=20, width=90, state="disabled", yscrollcommand=scrollbar.set)
        self.log_box.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.log_box.yview)
        
        # Запускаем обработчик логов
        self.process_log_queue()

    def process_log_queue(self):
        """Обрабатываем сообщения из очереди логов"""
        try:
            while True:
                message = self.log_queue.get_nowait()
                self.log_box.config(state="normal")
                self.log_box.insert(tk.END, f"{message}\n")
                self.log_box.see(tk.END)
                self.log_box.config(state="disabled")
        except:
            pass
        finally:
            self.root.after(100, self.process_log_queue)

    def log(self, message):
        """Добавляем сообщение в очередь логов"""
        self.log_queue.put(message)

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.file_path = path
            self.file_label.config(text=os.path.basename(path), fg="black")
        else:
            self.file_label.config(text="Файл не выбран", fg="gray")

    def test_api(self):
        if not self.account_var.get():
            messagebox.showwarning("Ошибка", "Выберите магазин.")
            return
        account_name = self.account_var.get()
        account = next(acc for acc in self.config["accounts"] if acc["name"] == account_name)
        Thread(target=self.test_api_connection, args=(account,), daemon=True).start()

    def test_api_connection(self, account):
        try:
            self.log("=== ТЕСТ API СОЕДИНЕНИЯ ===")
            headers = get_headers(account)
            
            # Тестируем простой запрос
            test_url = f"{API_BASE}/v3/product/info/list"
            test_payload = {"offer_id": ["test"], "product_id": []}
            
            self.log(f"Тестируем соединение с {test_url}")
            
            resp = requests.post(test_url, headers=headers, json=test_payload, timeout=10)
            
            self.log(f"Статус ответа: {resp.status_code}")
            
            if resp.status_code == 200:
                self.log("✅ API работает корректно")
            elif resp.status_code == 401:
                self.log("❌ Ошибка авторизации - проверьте Client-Id и Api-Key")
            elif resp.status_code == 403:
                self.log("❌ Доступ запрещен - проверьте права API ключа")
            else:
                self.log(f"⚠️ Неожиданный статус: {resp.status_code}")
                
        except Exception as e:
            self.log(f"❌ Ошибка подключения: {e}")

    def run_process(self):
        if not self.file_path:
            messagebox.showwarning("Ошибка", "Сначала выберите файл товаров.")
            return
        if not self.account_var.get():
            messagebox.showwarning("Ошибка", "Выберите магазин.")
            return
        account_name = self.account_var.get()
        account = next(acc for acc in self.config["accounts"] if acc["name"] == account_name)
        Thread(target=self.process, args=(account,), daemon=True).start()

    def process(self, account):
        try:
            self.progress['value'] = 0
            self.log("=== НАЧАЛО ОБРАБОТКИ ===")
            self.log("Загружаем файл...")
            
            offer_ids = read_offer_ids_xlsx(self.file_path)
            self.log(f"Найдено {len(offer_ids)} товаров для проверки.")
            
            if not offer_ids:
                self.root.after(0, lambda: messagebox.showinfo("Инфо", "В файле не найдено offer_id."))
                return

            self.progress['maximum'] = len(offer_ids) + 10
            self.progress['value'] = 10

            # Получаем информацию о товарах
            products, api_errors = get_products_info(account, offer_ids, self.log)
            self.progress['value'] = 50

            if not products and api_errors:
                self.log("❌ Не удалось получить данные о товарах")
                for error in api_errors[:5]:
                    self.log(f"Ошибка: {error}")
                return

            self.log(f"Получено данных о {len(products)} товарах")

            # Анализируем товары
            to_unarchive = []
            report_data = []

            for p in products:
                offer_id = p.get("offer_id", "unknown")
                product_id = p.get("id")
                is_autoarchived = p.get("is_autoarchived", False)
                is_archived = p.get("is_archived", False)  # Исправлено: is_archived вместо archived
                statuses = p.get("statuses", {})
                status = statuses.get("status_name", "unknown") if isinstance(statuses, dict) else "unknown"

                row_data = {
                    "offer_id": offer_id,
                    "product_id": product_id,
                    "status_before": status,
                    "archived": is_archived,
                    "is_autoarchived": is_autoarchived,
                    "action": "SKIPPED",
                    "error": ""
                }

                # Исправлена логика проверки
                if is_archived and not is_autoarchived:
                    to_unarchive.append(product_id)
                    row_data["action"] = "TO_UNARCHIVE"
                    self.log(f"{offer_id} → Будет восстановлен")
                elif is_archived and is_autoarchived:
                    row_data["action"] = "SKIPPED (AUTOARCHIVED)"
                    self.log(f"{offer_id} → Пропущен (автоархивирован)")
                elif not is_archived:
                    row_data["action"] = "SKIPPED (NOT_ARCHIVED)"
                    self.log(f"{offer_id} → Пропущен (не в архиве)")

                report_data.append(row_data)

            self.progress['value'] = 70

            # Восстанавливаем товары
            unarchive_results = []
            unarchive_errors = []
            
            if to_unarchive:
                self.log(f"Восстанавливаем {len(to_unarchive)} товаров...")
                unarchive_results, unarchive_errors = unarchive_products(account, to_unarchive, self.log)
                
                if unarchive_errors:
                    self.log(f"Ошибки при восстановлении: {len(unarchive_errors)}")
            else:
                self.log("Нет товаров для восстановления.")

            self.progress['value'] = 90

            # Добавляем информацию об отсутствующих товарах
            found_offer_ids = {p.get("offer_id") for p in products}
            for offer_id in offer_ids:
                if offer_id not in found_offer_ids:
                    report_data.append({
                        "offer_id": offer_id,
                        "product_id": "NOT_FOUND",
                        "status_before": "NOT_FOUND",
                        "archived": "NOT_FOUND",
                        "is_autoarchived": "NOT_FOUND",
                        "action": "NOT_FOUND",
                        "error": "Товар не найден в API"
                    })

            # Сохраняем отчёт
            out_path = os.path.join(os.path.dirname(self.file_path), "result.xlsx")
            write_report_xlsx(out_path, report_data)
            
            self.progress['value'] = 100
            
            self.log("=== ЗАВЕРШЕНО ===")
            self.log(f"Всего товаров обработано: {len(offer_ids)}")
            self.log(f"Найдено в API: {len(products)}")
            self.log(f"Восстановлено: {len(to_unarchive)}")
            self.log(f"Результат сохранён в {out_path}")
            
            # Показываем сообщение в основном потоке
            self.root.after(0, lambda: messagebox.showinfo(
                "Готово!", 
                f"Обработка завершена!\nВосстановлено: {len(to_unarchive)} товаров\nОтчёт: {out_path}"
            ))

        except Exception as e:
            error_msg = f"Критическая ошибка: {str(e)}"
            self.log(error_msg)
            self.root.after(0, lambda: messagebox.showerror("Ошибка", str(e)))


if __name__ == "__main__":
    root = tk.Tk()
    app = OzonUnarchiverApp(root)
    root.mainloop()
